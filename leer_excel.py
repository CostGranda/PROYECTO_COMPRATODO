from openpyxl import load_workbook as lw
import random as rnd
import string

def placa():
    letras = ''.join(rnd.choice(string.ascii_uppercase) for _ in range(3))
    numeros = ''.join(rnd.choice(string.digits) for _ in range(3))
    return letras+numeros

def documentos():
    return str(rnd.randint(1042, 7704))
    
## Leer archivo
archivo = lw('Malla_Ing_Informatica.xlsx', read_only=True)
hoja = archivo.getget_sheet_by_name('Hoja3')
nombre_tabla= ' Hoja3 '
cant_filas = hoja.max_row
cant_columnas = hoja.max_column
#####

def generar_inserts(cant_filas, hoja):
    lista_valores = []
    lst_sql = []
    k=0
    for i in range(1,cant_filas):
        lista_valores.append(hoja.cell(row=i, column=2).value)    
        lst_sql.append('INSERT INTO' + nombre_tabla + 
                    'VALUES (seq_codigos.nextval, '+ placa() +', '+lista_valores[k]+', '+documentos()+');')
        k+=1
    return lst_sql

INSERTS = generar_inserts(cant_filas, hoja)

#
#generar_insert(can)    
#cedula = random.randint(262309,1042770)
#
#
#for i in lista_valores:
#    lst_sql.append('INSERT INTO' + nombre_tabla + 
#                    'VALUES (seq_codigos.nextval, '+ placa() +', '+i+');')
    